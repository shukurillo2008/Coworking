from django.http import HttpResponse, JsonResponse
from django.shortcuts import render,redirect
from django.contrib import messages
from . import models
from django.utils import timezone
from decimal import Decimal 
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.db.models import Q
import xlrd
import openpyxl

@login_required(login_url='login_url')
def index(request):
    student = models.Student.objects.all()
    visit = models.EnterExit.objects.all()
    worker = User.objects.filter(is_superuser=False)
    visits_all = 0
    visits_today = 0
    used_degree = 0
    used_degree_all = 0
    not_used_degree = 0
    students = student.count()
    visits_out = models.OnOfTime.objects.filter(on_time__date=timezone.now().date()).count()
    money = 0
    all_money = 0

    for degree in models.UsedDegree.objects.all():
        used_degree_all += degree.used_degree
        if degree.created_time.day == timezone.now().day and degree.created_time.year == timezone.now().year:
            used_degree += degree.used_degree

    for degree in student:
        not_used_degree += degree.degree

    for visit in visit:
        visits_all += 1
        if visit.enter_time.day == timezone.now().day and visit.enter_time.year == timezone.now().year:
            visits_today += 1

    for i in models.Money.objects.all():
        all_money += i.money
        if i.time.on_time.date() == timezone.now().date():
            money += i.money 

    context = {
        'students':students,
        'visits_today':visits_today,
        'visits_all':visits_all,
        'used_degree':used_degree,
        'used_degree_all':used_degree_all,
        'not_used_degree':not_used_degree,
        'workers':worker,
        'price':models.TimePrice.objects.last(),
        'all_money':all_money,
        'money':money,
        'visits_out':visits_out
    }

    return render(request, 'dashboard.html', context)


def student_list(request):
    search_student = request.GET.get('search')
    students = models.Student.objects.all()

    if search_student:
        if search_student.isdigit():
            students = students.filter(origin_id=search_student)
        else:
            students = students.filter(
                Q(first_name__icontains=search_student) | Q(last_name__icontains=search_student)
            ).order_by('-degree')
            return render(request, 'student_list.html', {'page' : students})

    else:
        students = students.filter(is_student = True).order_by('-degree') 

    items_per_page = 10 
    paginator = Paginator(students, items_per_page)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    page_numbers = []

    if page.number >= 3: 
        page_numbers.append(1)
        page_numbers.append('...')

    for num in range(max(page.number - 2, 1), min(page.number + 3, paginator.num_pages + 1)):
        page_numbers.append(num)

    if page.number <= paginator.num_pages - 2: 
        page_numbers.append('...')
        page_numbers.append(paginator.num_pages)

    context = {
        'page': page,
        'page_numbers': page_numbers
    }
    
    return render(request, 'student_list.html', context)


def outed_student_list(request):
    students = models.Student.objects.filter(is_student=False)
    items_per_page = 10 
    paginator = Paginator(students, items_per_page)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    page_numbers = []

    if page.number >= 3: 
        page_numbers.append(1)
        page_numbers.append('...')

    for num in range(max(page.number - 2, 1), min(page.number + 3, paginator.num_pages + 1)):
        page_numbers.append(num)

    if page.number <= paginator.num_pages - 2: 
        page_numbers.append('...')
        page_numbers.append(paginator.num_pages)

    context = {
        'page': page,
        'page_numbers': page_numbers
    }

    return render(request, 'outed_student_list.html', context)


@login_required(login_url='login_url')
def en_ex_list(request):
    enter_exit = models.EnterExit.objects
    from_time = request.GET.get('from_time')
    to_time = request.GET.get('to_time')
    if from_time:
        if to_time:
            enter_exit = enter_exit.filter(enter_time__date__gte = from_time, enter_time__date__lte = to_time).order_by('-enter_time')
        else:
            enter_exit = enter_exit.filter(enter_time__gte = from_time).order_by('-enter_time')
    elif to_time:
        enter_exit = enter_exit.filter(enter_time__lte = to_time)
    else:
        enter_exit = enter_exit.all().order_by('-enter_time')
    items_per_page = 10
    paginator = Paginator(enter_exit, items_per_page)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    page_numbers = []

    if page.number >= 3: 
        page_numbers.append(1)
        page_numbers.append('...')

    for num in range(max(page.number - 2, 1), min(page.number + 3, paginator.num_pages + 1)):
        page_numbers.append(num)

    if page.number <= paginator.num_pages - 2: 
        page_numbers.append('...')
        page_numbers.append(paginator.num_pages)

    context = {
        'page': page,
        'page_numbers': page_numbers
    }
    return render(request, 'student_en_ex_list.html', context)


@login_required(login_url='login_url')
def create_student(request):
    if request.method == 'POST':
        try:
            first_name = request.POST['first_name']
            last_name = request.POST['last_name']
            origin_id = request.POST['origin_id']

            # qr_image = qrcode.make(origin_id)
            # image_buffer = io.BytesIO()
            # qr_image.save(image_buffer, format='PNG')

            student = models.Student.objects.create(
                first_name=first_name,
                last_name=last_name,
                origin_id=origin_id
            )


            # student.qr_code.save(f'qr_{origin_id}.png', ContentFile(image_buffer.getvalue()), save=True)
            messages.success(request, 'Created successfully!')
        except: 
            messages.warning(request, 'This ID is Used or Some thing went wrong =(')
        return redirect('students_url')
    else:
        return render(request, 'students_create.html')


@login_required(login_url='login_url')
def create_student_by_file(request):
    file = request.FILES.get("file")
    f = models.Files.objects.create(file=file)
    
    if f.file.name.endswith('xls'):
        book = xlrd.open_workbook(file_contents=f.file.read())
        sh = book.sheet_by_index(0)

        for row_num in range(1, sh.nrows):

            row = sh.row_values(row_num)
            origin_id = row[0]
            first_name = row[1]
            last_name = row[2]
            degree = row[3]
            is_student = row[4]
            
            try:
                student = models.Student.objects.get(origin_id=origin_id)
                student.first_name = first_name
                student.last_name = last_name
                if degree:
                    student.degree = degree
                if is_student == "+":
                    student.is_student = True
                elif is_student == "-":
                    student.is_student = False
                student.save()
            except:
                if origin_id and first_name and last_name: 
                    models.Student.objects.create(first_name = first_name, last_name = last_name, origin_id = origin_id, degree = degree)
        messages.success(request, 'done')

    elif f.file.name.endswith('.xlsx'):
        book = openpyxl.load_workbook(f.file)
        sh = list(book.worksheets[0].iter_rows(values_only=True))
        
        for up, row in enumerate(sh):
            if up == 0:
                continue
            
            origin_id = row[0]
            
            if isinstance(origin_id, int):
                try:
                    student = models.Student.objects.get(origin_id=origin_id)
                    student.first_name = row[1]
                    student.last_name = row[2]
                    if row[3] is not None: 
                        student.degree = row[3]

                    if row[4] == '+':
                        student.is_student = True
                    elif row[4] == '-':
                        student.is_student = False
                    student.save()
                except:
                    if row[3] is not None and row[1] is not None and row[2] is not None:
                        models.Student.objects.create(first_name=row[1], last_name=row[2], origin_id=origin_id, degree=row[3])

        messages.success(request, 'done')
    else:
        messages.error(request, 'file format must be xlsx or xls')
                        
    return redirect("students_url")


@login_required(login_url='login_url')
def student_table_make(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
    students = models.Student.objects.filter(is_student = True).order_by('id')
    wb = openpyxl.Workbook()
    ws = wb.active

    columns = ['Origin ID', 'First Name', 'Last Name', 'Degree', 'IS Student']
    ws.append(columns)

    for student in students:
        row = [student.origin_id, student.first_name, student.last_name, str(student.degree), '+' if student.is_student == True else "-"]
        ws.append(row)

    wb.save(response)
    return response


@login_required(login_url='login_url')
def outed_student_table_make(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="outed-students.xlsx"'
    students = models.Student.objects.filter(is_student = False).order_by('id')
    wb = openpyxl.Workbook()
    ws = wb.active

    columns = ['Origin ID', 'First Name', 'Last Name', 'Degree', 'IS Student']
    ws.append(columns)

    for student in students:
        row = [student.origin_id, student.first_name, student.last_name, str(student.degree), '+' if student.is_student == True else "-"]
        ws.append(row)

    wb.save(response)
    return response


@login_required(login_url='login_url')
def student_edit(request, id):
    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        try:
            student = models.Student.objects.get(origin_id=id)
            student.first_name = first_name
            student.last_name = last_name
            student.save()
            messages.success(request , 'Changed successfuly!')
        except:
            messages.warning(request, 'Some thing went wrong =(')
        return redirect('student_edit_url',id)
    else:
        student = models.Student.objects.get(origin_id=id)
        used_degree = models.UsedDegree.objects.filter(student=student).order_by('-id')
        items_per_page = 10  
        paginator = Paginator(used_degree, items_per_page)
        page_number = request.GET.get('page')
        page = paginator.get_page(page_number)

        page_numbers = []

        if page.number >= 3: 
            page_numbers.append(1)
            page_numbers.append('...')

        for num in range(max(page.number - 2, 1), min(page.number + 3, paginator.num_pages + 1)):
            page_numbers.append(num)

        if page.number <= paginator.num_pages - 2: 
            page_numbers.append('...')
            page_numbers.append(paginator.num_pages)
        context = {
            'student': student,
            'degree': page,
            'page_numbers':page_numbers
        }

    return render(request, 'student_detail.html', context)


@login_required(login_url='login_url')
def student_delete(request):
    origin_id = request.POST.get('origin_id')
    student = models.Student.objects.get(origin_id=origin_id)
    student.delete()
    messages.success(request, 'Deleted successfully')

    return redirect('students_url')


def enter_exit(request):
    context = {}

    if request.method == 'POST':
        origin_id = request.POST.get('user_id')
        if origin_id:
            try:
                student = models.Student.objects.get(origin_id=origin_id)
                if student.is_student:
                    try:
                        enterexit = models.EnterExit.objects.get(student=student, in_out=False)
                        price = models.TimePrice.objects.last()
                        enterexit.in_out = True
                        enterexit.exit_time = timezone.now()
                        enterexit.save()
                        
                        time_difference = enterexit.exit_time - enterexit.enter_time
                        time_hours = time_difference.total_seconds() / 3600
                        
                        total_cost = price.price * Decimal(str(time_hours))
                        student.degree -= total_cost

                        send_degree = None

                        if student.degree < 0:
                            send_degree = student.degree

                        models.UsedDegree.objects.create(
                            student = student,
                            enter_exit = enterexit,
                            before_degree = models.Student.objects.get(origin_id=origin_id).degree,
                            used_degree = total_cost,
                            after_degree = student.degree,
                        )

                        student.save()
                        
                        context['student'] = student
                        return JsonResponse({
                            'last_name': student.last_name,
                            "first_name": student.first_name,
                            'status': "Chiqdi",
                            'degree':  f'{total_cost:.2f}',
                            'send_degree': send_degree
                        })
                    except:

                        status = ''
                        if student.degree >= models.TimePrice.objects.last().price:
                            models.EnterExit.objects.create(
                                student = student,
                                enter_time = timezone.now(),
                                in_out = False,
                            )
                            status = 'Kirdi'
                        else:
                            status = 'yetarli bal mavjud emas' 
                        return JsonResponse({
                            'last_name': student.last_name,
                            "first_name": student.first_name,
                            "status": status,
                        })
                else:
                    # messages.error(request, 'You are not student anymore')
                    return JsonResponse({
                            "status": 'You are not student anymore',
                        })
            except:
                messages.error(request, 'Not Found 404')
                return redirect('error_url')
            
    return render(request, 'enter_exit.html', context)


def error(request):
    return render(request, 'error.html')


@login_required(login_url='login_url')
def change_price(request):
    try:
        price = request.POST.get('price')
        price_time = models.TimePrice.objects.last()
        price_time.price = price
        price_time.save()
        messages.success(request, 'Saved successfully!')
        return redirect('index_url')
    except:
        models.TimePrice.objects.create(
            price = request.POST.get('price'),
        )
        messages.success(request, 'created')
        return redirect('index_url')
    

@login_required(login_url='login_url')
def add_degree(request):
    try:
        degree = Decimal(request.POST.get('degree'))
        student_id = int(request.POST.get('origin_id'))

        student = models.Student.objects.get(origin_id=student_id)
        old_degree = student.degree
        new_degree = old_degree + degree

        added_degree = models.AddedDegree.objects.create( 
            student=student,
            before_degree=old_degree,
            added_degree=degree, 
            after_degree=new_degree
        )

        student.degree = new_degree
        student.save()
        messages.success(request, 'Added successfully!')
    except:
        messages.warning(request, 'NOt added, some thing went wrong =(') 
    return redirect('student_edit_url',student_id)


@login_required(login_url='login_url')
def change_components(request):
    company = models.CompanyComponent.objects.last()
    if request.method == 'POST':
        try:
            company_name = request.POST.get('company_name')
            about = request.POST.get('about')
            logo = request.FILES.get('logo')
            company.company_name = company_name
            company.about = about            
            if logo :
                company.logo = logo
            company.save()
        except:
            models.CompanyComponent.objects.create(
                about =about,
                logo = logo,
                company_name = company_name
            )

    context = {
        'price':models.TimePrice.objects.last(),
        'company':company,
        'price_for_students': models.TimeMoney.objects.get(for_student=True),
        'price_for_strangers': models.TimeMoney.objects.get(for_student=False)
    }

    return render(request, 'change_info.html', context)


@login_required(login_url='login_url')
def change_price_money(request):
    price = request.POST.get('price')
    for_student = request.POST.get('for_student')
    try:
        if for_student == 'true':
            price_time = models.TimeMoney.objects.get(for_student = True)
        else:
            price_time = models.TimeMoney.objects.get(for_student = False)
        price_time.price = price
        price_time.save()
        messages.success(request, 'Saved successfully!')
        return redirect('index_url')
    except:
        if for_student == 'true':
            type_price = True
        else:
            type_price = False
        models.TimeMoney.objects.create(
            price = price,
            for_student = type_price
        )
        messages.success(request, 'created')
        return redirect('index_url')


@login_required(login_url='login_url')
def company(request):
    company = models.CompanyComponent.objects.last()
    return {'objects': company}


@login_required(login_url='login_url')
def pc_list(request):
    pc = models.Pc.objects.all().order_by('number')
    context = {
        "pc": pc
    }

    return render(request, 'pc_list.html', context)


@login_required(login_url='login_url')
def pc_st_end(request):
    pc_number = request.POST.get('pc_number')
    is_student = request.POST.get('is_student')
    
    try:
        pc = models.Pc.objects.get(number=int(pc_number))
        if is_student == 'true':
            time_price = models.TimeMoney.objects.get(for_student=True)
        else:
            time_price = models.TimeMoney.objects.get(for_student=False)     
        try:
            on_of = models.OnOfTime.objects.get(pc=pc, off_time__isnull=True)
            on_of.off_time = timezone.now()
            on_of.save()
            time_difference = on_of.off_time - on_of.on_time
            time_hours = time_difference.total_seconds() / 3600
            total_cost = on_of.pay_for.price * Decimal(time_hours)
            models.Money.objects.create(pc=pc, time=on_of, money=total_cost)
            pc.status = 2
            pc.save()
            messages.success(request, f'costs : {round(total_cost, 2)}')
            return redirect('pc_list_url')
        except:
            on_of = models.OnOfTime.objects.create(
                pc=pc,
                on_time=timezone.now(), 
                pay_for=time_price
                )
            pc.status = 1
            pc.save()
            messages.success(request, 'Started')
            return redirect('pc_list_url')
    except:

        messages.error(request, 'something went wrong  =/')
        return redirect('error_url')


@login_required(login_url='login_url')
def add_pc(request):
    number = request.POST.get('pc_number')
    models.Pc.objects.create(number=number)
    messages.success(request, 'Added')
    return redirect('pc_list_url')


@login_required(login_url='login_url')
def delete_pc(request):
    number = request.POST.get('pc_number')
    models.Pc.objects.get(number=number).delete()

    return redirect('pc_list_url')


def edit_pc(request):
    number = request.POST.get('pc_number')
    pc = models.Pc.objects.get(number=number)
    if pc.status == 3:
        pc.status = 2
        pc.save()
    else:
        pc.status = 3
        pc.save()
    messages.success(request, 'Done')
    return redirect('pc_list_url')










